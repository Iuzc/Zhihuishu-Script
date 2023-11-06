import sqlite3
import os
import openpyxl
#连接数据库
def dbHandle():
    '''
    用于创建并返回与 SQLite3 数据库的连接对象
    :return:
    '''
    #创建一个与名为 "Course.db" 的 SQLite3 数据库文件的连接。如果该数据库文件不存在，则会创建一个新的数据库文件
    conn = sqlite3.connect("Course.db")
    #将数据库连接的事务隔离级别设置为 "EXCLUSIVE"，也就是在一个事务进行时，其他事务无法同时进行
    conn.isolation_level = 'EXCLUSIVE'
    return conn
class Course():
    # 数据库连接池
    conn = None
    # 游标，表示数据库语句
    cursor = None
    # 各个用户的数据库唯一标识
    name = ""
    def __init__(self,name=""):
        self.conn = dbHandle()
        self.name = name
        self.cursor = self.conn.cursor()
        sql = "create table if not exists " + self.name +\
            "(id int auto_increment primary key," \
            "name varchar(50) null," \
            "isStudies int null);"
        if name != "":
            self.cursor.execute(sql)
    def addItem(self, item):
        try:
            self.cursor.execute("INSERT INTO "+ self.name+" values(?,?,?) ",(item["index"],item["name"],item["isStudies"]))
            self.conn.commit()
            print(item["name"]+" 成功插入数据库")
        except Exception as e:
            print(e)
            self.conn.rollback()
    def findById(self, id):
        sql = "select * FROM "+self.name +" where id = " + str(id)
        # 执行SQL语句
        self.cursor.execute(sql)
        result = self.cursor.fetchone()
        # 中文转码
        result = list(result)
        result_dic = {
            "id" : result[0],
            "name":result[1],
            "isStudies":result[2]
        }
        return result_dic
    def updataById(self, id):
        sql = "update "+ self.name + " set isStudies = 1 where id = " + str(id)
        # 执行SQL语句
        cursor = self.cursor
        cursor.execute(sql)
        self.conn.commit()
        print("修改" + str(id) + "号数据成功")
    def updataByTitle(self, title):
        sql = "update "+ self.name + " set isStudies = 1 where name = ' " + str(title) + "'"
        # 执行SQL语句
        cursor = self.cursor
        cursor.execute(sql)
        self.conn.commit()
        print("修改" + str(title) + "数据成功")
    def selectAll(self):
        sql = "SELECT * FROM " +self.name + " ORDER BY id ASC"
        cursor = self.cursor
        cursor.execute(sql)
        rows = cursor.fetchall()
        # 将查询结果转换为列表
        column_names = [description[0] for description in cursor.description]
        # 将查询结果转换为字典列表
        result = []
        for row in rows:
            row_dict = {}
            for i in range(len(column_names)):
                row_dict[column_names[i]] = row[i]
            result.append(row_dict)
        return result
    def checkTableEmpty(self):
        '''
        判断用户表是否为空
        :return:
        '''
        cursor = self.cursor
        sql = f"SELECT COUNT(*) FROM {self.name} "
        cursor.execute(sql)
        result = cursor.fetchone()

        return result[0] == 0
    def destroy(self):
        self.conn.close()
        print("清除课程数据中...")
        os.remove('Course.db')
    def outputExcel(self):
        # 从数据库中读取数据
        cursor = self.cursor
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        workbook = openpyxl.Workbook()
        # 遍历所有表
        for table_name in tables:
            # 创建一个工作表
            worksheet = workbook.create_sheet(table_name[0])
            cursor.execute(f"SELECT * FROM {table_name[0]} ORDER BY id ASC")
            rows = cursor.fetchall()
            # 写入表头
            for i, column in enumerate(cursor.description):
                worksheet.cell(row=1, column=i + 1, value=column[0])
            # 写入数据
            for r, row in enumerate(rows, start=2):
                for c, cell in enumerate(row):
                    worksheet.cell(row=r, column=c + 1, value=cell)
        # 保存Excel文件
        workbook.save('Course.xlsx')
        print("\n答案表已导出至当前文件夹")
