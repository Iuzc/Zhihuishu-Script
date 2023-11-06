import os
import sqlite3
import openpyxl
#连接数据库
def dbHandle():
    # sqlite3写法
    conn = sqlite3.connect("Answer.db")
    # conn.isolation_level = 'EXCLUSIVE'
    return conn

class Answer(object):
    # 数据库连接池
    conn = None
    # 游标，表示数据库语句
    cursor = None
    # 各个用户的数据库唯一标识
    name = ""
    def __init__(self,name=""):
        self.conn = dbHandle()
        self.name = name
        self.cursor = self.conn.cursor()
        # sql = "drop table if exists " + self.name
        # self.cursor.execute(sql)
        sql = "create table if not exists " + self.name + \
              "(title varchar(100) null," \
              "answer varchar(100) null," \
              "chapter varchar(10) null);"
        if name != "":
            self.cursor.execute(sql)
    def addItem(self, item):
        try:
            self.cursor.execute("INSERT INTO "+ self.name+" values(?,?,?) ",(item["title"],item["answer"],item["chapter"]))
            self.conn.commit()
            print(item["title"]+"插入成功")
        except Exception as e:
            print(e)
            self.conn.rollback()
        # 谨记下面注释不可取消，增加的时候不能关
        # self.cursor.close()
        # self.conn.close()
    def findByTitle(self, title):
        sql = "select answer FROM "+self.name +" where title = " + title
        print(sql)
        # 执行SQL语句
        self.cursor.execute(sql)
        result = self.cursor.fetchone()
        # 中文转码
        result = list(result)
        print("查询" + title + "号数据成功")
        # self.cursor.close()
        # self.conn.close()
        return result
    def selectAll(self, chapter_name):
        sql = "SELECT answer FROM " +self.name +" where chapter = '" + chapter_name + "'"
        print(sql)
        cursor = self.cursor
        cursor.execute(sql)
        rows = cursor.fetchall()
        # 将查询结果转换为列表
        column_names = [description[0] for description in cursor.description]
        # 将查询结果转换为字典列表
        result = []
        for row in rows:
            for i in range(len(column_names)):
                result.append(row[i])
        return result
    def checkTableEmpty(self):
        cursor = self.cursor
        sql = f"SELECT COUNT(*) FROM {self.name}"
        cursor.execute(sql)
        result = cursor.fetchone()
        # 获取查询结果
        table_empty = result[0] == 0
        print(table_empty)
        return table_empty
    def destroy(self):
        self.conn.close()
        print("清除答案数据中...")
        os.remove('Answer.db')
    def outputExcel(self):
        # 从数据库中读取数据
        cursor = self.cursor
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        print(tables)
        workbook = openpyxl.Workbook()
        # 遍历所有表
        for table_name in tables:
            # 创建一个工作表
            worksheet = workbook.create_sheet(table_name[0])
            cursor.execute(f"SELECT * FROM {table_name[0]}")
            rows = cursor.fetchall()

            # 写入表头
            for i, column in enumerate(cursor.description):
                worksheet.cell(row=1, column=i + 1, value=column[0])
            # 写入数据
            for r, row in enumerate(rows, start=2):
                for c, cell in enumerate(row):
                    worksheet.cell(row=r, column=c + 1, value=cell)
        # 保存Excel文件
        workbook.save('Answers.xlsx')
        print("\n答案表已导出至当前文件夹")
if __name__ == "__main__":
    db_answer=Answer()
    db_answer.outputExcel()